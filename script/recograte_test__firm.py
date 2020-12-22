import serial
import time
import datetime
import winsound
import os
from openpyxl import load_workbook
import re
#from common.re_num_to_characters import ReplaceNumCharacters


# 全局变量-文件配置
SERIAL_NAME="COM3"  # 串口的名称
WAVE_RATE=115200   # 波特率
WAKEUP_WORD="mcrc_wk_cb"  # 唤醒标志

# 环境配置
wakeup_path = "D:/思必驰方言/canton-out/04_叶雄斌_000.wav"  #唤醒文件：你好小美的音频文件
root_path = "D:/思必驰方言/canton-out/"  # 需要执行的，音频文件的根目录

LOG_OUTPUT_DIRECTORY = 'D:/log2/'  # 日志文件输出目录
logname="粤语左45度-5M-噪声_20200417"  # 日志名称
excel_path = 'D:/思必驰方言/讯飞/粤语/10人粤语.xlsx'  # 映射关系表，命令词
now = time.strftime('%Y-%m-%d-%H-%M-%S')
result_excel_path = 'D:/思必驰方言/讯飞/粤语/测试结果/10人粤语_左45度-5M-噪声_%s.xlsx'%now  # 测试结果

run_lanuage=""   # 普通话：为空， 粤语：cantonese，四川话：sichuanese

startRows=12 #开始行
run_method = "all"  # 执行方式，执行奇数：odd_number； ---还是偶数：even_number； ---所有行都执行：all。




def write_file(fp, lines, mode='a+'):
    with open(fp, mode,encoding="utf-8") as f:
        f.writelines(lines)
        f.writelines("\n")
        f.close()


def recvWakeup(serial,logfd):
    i = 0
    # 5秒内收到数据
    new_data=""
    while i < 10:
        print("接收信息.....")
        data = serial.read_all()
        print(data)
        if len(data) == 0:
            # continue
            time.sleep(0.1)
            i += 1
        elif isinstance(data,bytes):
            print(data)
            new_data=new_data+data.decode(encoding='utf-8', errors='ignore')
            print(new_data)
            log_ts_str = "[" + datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S.%f') + "] " + new_data
            write_file(logfd, log_ts_str)
            if new_data.find(WAKEUP_WORD)!=-1:
                break
            #break

    ret=new_data.find(WAKEUP_WORD)
    if ret != -1:
        print('wakeup success')
        return True,""
    else:
        print('wakeup failure')
        return False,""


# 唤醒小美,读取音频文件
def wakeup_File(serialFd,log_filename):
    # 定义全局变量
    global totalWk, cntWkSucc, cntWkFail

    write_file(log_filename, "播放唤醒命令词")
    winsound.PlaySound(wakeup_path, winsound.SND_FILENAME)  # 播放唤醒词
    totalWk = totalWk + 1
    is_wakeup, data = recvWakeup(serialFd,log_filename)  # 接收是否唤醒成功
    # 是否唤醒
    print("是否唤醒: " + str(is_wakeup))
    if is_wakeup == True:
        cntWkSucc = cntWkSucc + 1
        return True

    else:
        cntWkFail = cntWkFail + 1
        iswake = wakeup_File(serialFd,log_filename)
        return iswake

def getasrvalue(newdata):
    try:
        if newdata.find("TTS:") == -1:
            return "",""

        commond = re.compile(".*TTS.*text\"\:\"(.*)\"\,\"endSession\".*")
        c = commond.findall(newdata)
        print(c)
        asrvaluetxt = ""
        lanuage_value = ""
        if len(c) > 0:
            asrvaluetxt = str(c[0]).replace(" ", "")

        commond = re.compile(".*TTS.*url\"\:\"(.*)\"\,\"seq\".*")
        c = commond.findall(newdata)
        if len(c) > 0:
            lanuage_value = str(c[0]).replace(" ", "")

        return asrvaluetxt, lanuage_value

    except Exception as e:
        print("解析异常了：", e)
        return "", ""

    finally:
        pass


def recvCmd(serial, txt,log_filename):
    global cntCmdSucc, cntCmdFail, cntCmdTotal
    i = 0
    res = 'ASR结果 '
    # 10 秒内收到回应
    cmd_newdata = ""
    try:
        while i < 50:
            count = serial.inWaiting()
            if count > 0:
                time.sleep(0.1)
                buff = serial.read_all()
                if isinstance(buff, bytes):
                    temp_data = buff.decode(encoding='utf-8', errors='ignore')
                    cmd_newdata = cmd_newdata + temp_data

            else:
                time.sleep(0.1)
                i += 1
                # print "asr loop"
    except Exception as e:
        print("获取asr结果异常：" + str(e))
        write_file(log_filename,"获取asr结果异常：" + str(e))
        #fileObject.write("获取asr结果异常：" + str(e))

    if len(cmd_newdata) <= 1:
        cmd_newdata = '空数据'

    write_file(log_filename,cmd_newdata)
    cmd_newdata=cmd_newdata.replace("\n","").replace("\t","").replace("\r","")
    # 语种
    serial_value,lanuage_value, =getasrvalue(cmd_newdata)

    serial.flushInput()
    cntCmdTotal = cntCmdTotal + 1
    print("asrvalue",serial_value)
    print("语种:",lanuage_value)
    ret = serial_value.find(txt)
    if ret != -1:
        cntCmdSucc = cntCmdSucc + 1
        return True, serial_value,lanuage_value
    else:
        if txt == serial_value:
            cntCmdSucc = cntCmdSucc + 1
            return True, serial_value, lanuage_value
        else:
            cntCmdFail = cntCmdFail + 1
            return False, serial_value,lanuage_value

# 获取excel，计算识别率 打断成功率 唤醒信息
def read_excel():
    serialFd = serial.Serial(SERIAL_NAME, WAVE_RATE, timeout=60)  # 7686 的串口

    log_filename = LOG_OUTPUT_DIRECTORY + logname + datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S') + '.log'  #日志文件

    workbook = load_workbook(excel_path)
    names = workbook.sheetnames

    allcase_wk=allSucss=0  #唤醒
    allCmdtotal=allCmdSucess=0  # 案例
    for onename in names:
        print("当前sheet名称: " + onename)
        worksheet =workbook[onename]

        # 全局变量，唤醒信息
        global totalWk, cntWkSucc, cntWkFail
        totalWk=cntWkSucc=cntWkFail=0

        # 全局变量，案例信息
        global cntCmdTotal, cntCmdSucc, cntCmdFail, ratioCmd
        cntCmdTotal = 0  # 案例总数
        cntCmdSucc = 0  # 识别成功次数
        cntCmdFail = 0  # 识别失败次数
        ratioCmd = 0.0  # 识别率

        lanuage_succ=lanuage_fail=0
        for row in range(startRows, worksheet.max_row + 1):
            if run_method == "even_number" and row % 2 == 0:  # 偶数
                continue
            elif run_method == "odd_number" and row % 2 != 0:  # 奇数
                continue

            wav_value = worksheet.cell(row, 1).value  # wav文件值
            if wav_value.strip() == '':
                continue
            # 判断文件是否存在
            wav_path = root_path + wav_value  # wav 路径
            print("wav_path: " + str(wav_path))
            if (os.path.exists(wav_path) == False):
                print("文件不存在")
                continue

            issucess = wakeup_File(serialFd,log_filename)  # 唤醒小美
            if issucess:
                # step 2-2, 播放命令词，读取串口，判断是否正确识别
                time.sleep(1)  # 间隔2.
                except_value = str(worksheet.cell(row, 2).value).strip()  # 预期结果值
                print("播放命令词: " + str(except_value))
                write_file(log_filename, "播放命令词: " + str(except_value))
                print("播放命令词: " + str(except_value))
                write_file(log_filename,"播放命令词: " + str(except_value))
                winsound.PlaySound(wav_path, winsound.SND_FILENAME)
                ars_sucess, ars_result ,asr_lanuage= recvCmd(serialFd, str(except_value),log_filename)  # 接收asr返回结果
                if asr_lanuage.find(run_lanuage)!=-1:
                    lanuage_succ=lanuage_succ+1
                    is_lanuage="true"
                    asr_lanuage = run_lanuage
                else:
                    lanuage_fail=lanuage_fail+1
                    is_lanuage = "false"
                    asr_lanuage = "putonghua"


                print("asr返回结果信息:" + ars_result)
                print("识别结果:" + str(ars_sucess))

                print("唤醒总数: " + str(totalWk))
                print("唤醒成功数: " + str(cntWkSucc))
                print("唤醒失败数: " + str(cntWkFail))
                ratioWk = (cntWkSucc / totalWk) * 100
                ratioWk_value = "{:.2f}%".format(ratioWk)
                print("唤醒成功率: " + str(ratioWk_value))

                print("案例总数: " + str(cntCmdTotal))
                print("识别成功次数: " + str(cntCmdSucc))
                print("识别失败次数: " + str(cntCmdFail))
                ratioCmd = (cntCmdSucc / cntCmdTotal) * 100
                ratioCmd_value = "{:.2f}%".format(ratioCmd)
                print("识别成功率: " + str(ratioCmd_value))

                print("语种成功次数: " + str(lanuage_succ))
                print("语种失败次数: " + str(lanuage_fail))
                lan_ratioCmd = (lanuage_succ / cntCmdTotal) * 100
                lan_ratioCmd_value = "{:.2f}%".format(lan_ratioCmd)
                print("语种成功率: " + str(lan_ratioCmd_value))

                # 写入excel表
                worksheet.cell(1, 2).value = str(cntCmdTotal)  # 案例总数
                worksheet.cell(2, 2).value = str(cntWkSucc) #唤醒成功数
                worksheet.cell(3, 2).value = str(cntWkFail)  # 唤醒失败数
                worksheet.cell(4, 2).value = str(ratioWk_value)  # 唤醒成功率

                worksheet.cell(5, 2).value = str(cntCmdSucc)  # 案例成功数
                worksheet.cell(6, 2).value = str(cntCmdFail)  # 案例识别失败数
                worksheet.cell(7, 2).value = str(ratioCmd_value)  # 案例识别成功率
                worksheet.cell(8, 2).value = str(lanuage_succ)  # 语言成功数

                worksheet.cell(9, 2).value = str(lanuage_fail)  # 语言失败数
                worksheet.cell(10, 2).value = str(lan_ratioCmd_value)  # 语言成功率

                worksheet.cell(row, 3).value = ars_result   # asr 的值
                worksheet.cell(row, 4).value = ars_sucess    # asr 识别率
                worksheet.cell(row, 5).value = asr_lanuage    # 语种
                worksheet.cell(row, 6).value = is_lanuage  # 语种是否正确

                if str(ars_sucess).lower()=="true" and str(is_lanuage)=="true":
                    worksheet.cell(row, 7).value = "true"  # 语种是否正确
                else:
                    worksheet.cell(row, 7).value = "false"  # 语种是否正确
                workbook.save(result_excel_path)  # 保存


        allcase_wk=allcase_wk+totalWk
        allSucss=allSucss+cntWkSucc
        allCmdtotal=allCmdtotal+cntCmdTotal
        allCmdSucess=allCmdSucess+cntCmdSucc

    print("所有案例唤醒总数: " + str(allcase_wk))
    print("所有案例唤醒成功数: " + str(allSucss))
    all_ratioWk = (allSucss / allcase_wk) * 100
    all_ratioWk_value = "{:.2f}%".format(all_ratioWk)
    print("所有案例唤醒成功率: " + str(all_ratioWk_value))

    print("所有案例_案例总数: " + str(allCmdtotal))
    print("所有案例_识别成功次数: " + str(allCmdSucess))
    all_ratioCmd = (allCmdSucess / allCmdtotal) * 100
    all_ratioCmd_value = "{:.2f}%".format(all_ratioCmd)
    print("所有案例_识别成功率: " + str(all_ratioCmd_value))





if  __name__ == '__main__':
    read_excel()
